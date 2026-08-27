# -*- coding: utf-8 -*-
import base64
import io
import logging

from odoo import _, fields, models
from odoo.exceptions import UserError

from ..models import bom_rules

_logger = logging.getLogger(__name__)

CONFIDENCE_LEVELS = {"auto": bom_rules.AUTO, "review": bom_rules.REVIEW}


class BomImport(models.TransientModel):
    _name = "tectora.bom.import"
    _description = "Stuklijsten importeren"

    file = fields.Binary(string="Bestand (.xlsx)", required=True, attachment=False)
    filename = fields.Char(string="Bestandsnaam")
    sheet_name = fields.Char(
        string="Tabblad", help="Laat leeg om het eerste tabblad te nemen."
    )
    product_min = fields.Selection(
        [
            ("auto", "Alleen zekere producten (score ≥ 0,70)"),
            ("review", "Ook waarschijnlijke producten (score ≥ 0,45)"),
        ],
        string="Producten koppelen",
        default="auto",
        required=True,
        help="Het bestand bevat geen productcode, dus het product wordt op "
        "naam gezocht. Zeker = de naam komt (bijna) overeen.",
    )
    component_min = fields.Selection(
        [
            ("auto", "Alleen zekere componenten (score ≥ 0,70)"),
            ("review", "Ook waarschijnlijke componenten (score ≥ 0,45)"),
        ],
        string="Componenten koppelen",
        default="auto",
        required=True,
        help="Componenten waarvoor geen product gevonden wordt, komen niet in "
        "de stuklijst en staan in het verslag.",
    )
    uom_policy = fields.Selection(
        [
            ("base_only", "Alleen regels in een echte eenheid (veilig)"),
            ("convert", "Verpakkingen omrekenen naar de basiseenheid"),
            ("verbatim", "Aantallen letterlijk overnemen"),
        ],
        string="Verpakkingseenheden",
        default="base_only",
        required=True,
        help="De eenheidkolom in het bestand is een verpakking ('Doos 500 "
        "stuks', 'm per 3m element'), geen Odoo-eenheid, en de basis van het "
        "aantal wisselt per regel. Veilig: die regels worden overgeslagen en "
        "opgelijst. Omrekenen: aantal × de inhoud van de verpakking — een "
        "aanname, controleer het verslag. Letterlijk: aantal zoals het er "
        "staat, met de eenheid van het product.",
    )
    dry_run = fields.Boolean(
        string="Alleen analyseren",
        default=True,
        help="Bekijkt het bestand en toont wat er zou gebeuren, zonder "
        "stuklijsten aan te maken.",
    )
    state = fields.Selection(
        [("draft", "Concept"), ("done", "Uitgevoerd")], default="draft"
    )
    summary = fields.Text(string="Verslag", readonly=True)
    unmatched = fields.Text(string="Niet gekoppeld", readonly=True)

    # ------------------------------------------------------------------ reading
    def _read_rows(self):
        self.ensure_one()
        try:
            import openpyxl  # Odoo ships openpyxl (see requirements.txt)
        except ImportError:
            raise UserError(
                _("De Python-bibliotheek openpyxl is niet beschikbaar op de server.")
            )
        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(base64.b64decode(self.file)),
                data_only=True,
                read_only=True,
            )
        except Exception as error:
            raise UserError(
                _("Het bestand kon niet gelezen worden als .xlsx-bestand (%s).",
                  error)
            )
        name = (self.sheet_name or "").strip()
        if name and name not in workbook.sheetnames:
            raise UserError(
                _("Tabblad '%(sheet)s' niet gevonden. Beschikbaar: %(all)s",
                  sheet=name, all=", ".join(workbook.sheetnames))
            )
        sheet = workbook[name] if name else workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            raise UserError(_("Het tabblad bevat geen gegevens."))
        return rows[1:]  # the header row is fixed, columns are positional

    # ------------------------------------------------------------------ import
    def action_import(self):
        self.ensure_one()
        boms = bom_rules.parse_export(self._read_rows())
        if not boms:
            raise UserError(
                _("Geen stuklijsten gevonden. Verwachte kolommen: product, "
                  "reeks, referentie, stuklijstsoort, stuklijstregels, aantal, "
                  "onderdeel, maateenheid onderdeel.")
            )
        report = self.env["mrp.bom"]._tectora_import_boms(
            boms,
            {
                "product_min": CONFIDENCE_LEVELS[self.product_min],
                "component_min": CONFIDENCE_LEVELS[self.component_min],
                "uom_policy": self.uom_policy,
                "dry_run": self.dry_run,
            },
        )
        _logger.info("tectora_boms: wizard %s -> %s", self.filename, report)
        self.write({
            "state": "done",
            "summary": "\n".join(self._summary_lines(report)),
            "unmatched": "\n".join(self._unmatched_lines(report)) or False,
        })
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
            "context": self.env.context,
        }

    def _summary_lines(self, report):
        lines = [
            _("Bestand: %s", self.filename or "-"),
            _("%(boms)s stuklijsten met %(lines)s regels gelezen",
              boms=report["boms"], lines=report["lines"]),
            "",
            _("Producten: %(auto)s zeker, %(review)s waarschijnlijk, "
              "%(manual)s zonder match",
              auto=report["products"]["auto"],
              review=report["products"]["review"],
              manual=report["products"]["manual"]),
            _("Componenten: %(auto)s zeker, %(review)s waarschijnlijk, "
              "%(manual)s zonder match",
              auto=report["components"]["auto"],
              review=report["components"]["review"],
              manual=report["components"]["manual"]),
            "",
        ]
        if report["dry_run"]:
            lines.append(
                _("Analyse: %(boms)s stuklijsten met %(lines)s regels zouden "
                  "aangemaakt worden.",
                  boms=report["planned"], lines=report["planned_lines"])
            )
        else:
            lines.append(
                _("%(new)s stuklijsten aangemaakt, %(upd)s bijgewerkt, "
                  "%(lines)s regels",
                  new=report["created"], upd=report["updated"],
                  lines=report["planned_lines"])
            )
        if report["duplicates"]:
            lines.append(
                _("%s identieke stuklijsten in het bestand samengevoegd",
                  report["duplicates"])
            )
        if report["no_product"]:
            lines.append(
                _("%s stuklijsten overgeslagen: geen product gevonden",
                  len(report["no_product"]))
            )
        if report["no_component"]:
            lines.append(
                _("%s regels overgeslagen: component niet gevonden",
                  len(report["no_component"]))
            )
        if report["uom_skipped"]:
            lines.append(
                _("%s regels overgeslagen: aantal staat in een verpakkings"
                  "eenheid", len(report["uom_skipped"]))
            )
        if report["uom_converted"]:
            lines.append(
                _("%s regels omgerekend van verpakking naar basiseenheid — "
                  "controleer die hieronder", len(report["uom_converted"]))
            )
        if report.get("crowded"):
            lines.append("")
            lines.append(
                _("Let op: deze producten krijgen veel stuklijsten. Bij een "
                  "variantenreeks is dat normaal; anders zijn meerdere "
                  "producten uit het bestand op hetzelfde product beland:")
            )
            for code, count in report["crowded"][:10]:
                lines.append("   %s: %s stuklijsten" % (code, count))
        return lines

    def _unmatched_lines(self, report):
        """The queue of human decisions, heaviest first."""
        out = []
        if report["no_product"]:
            out.append(_("— Producten zonder match (stuklijst niet ingelezen) —"))
            for name, score in sorted(report["no_product"], key=lambda x: -x[1]):
                out.append("  %.2f  %s" % (score, name))
            out.append("")
        if report["no_component"]:
            counted = {}
            for component, score, parent in report["no_component"]:
                entry = counted.setdefault(component, [score, 0])
                entry[1] += 1
            out.append(_("— Componenten zonder match (aantal regels) —"))
            for component, (score, count) in sorted(
                counted.items(), key=lambda kv: -kv[1][1]
            ):
                out.append("  %4dx  %.2f  %s" % (count, score, component))
            out.append("")
        if report["uom_skipped"]:
            counted = {}
            for component, uom, qty, parent in report["uom_skipped"]:
                counted.setdefault((component, uom), 0)
                counted[(component, uom)] += 1
            out.append(_("— Overgeslagen verpakkingseenheden (aantal regels) —"))
            for (component, uom), count in sorted(
                counted.items(), key=lambda kv: -kv[1]
            ):
                out.append("  %4dx  %-30s %s" % (count, uom, component))
            out.append("")
        if report["uom_converted"]:
            out.append(_("— Omgerekend (aantal in het bestand → aantal in de "
                         "stuklijst) —"))
            seen = set()
            for component, uom, before, after in report["uom_converted"]:
                key = (component, uom, before)
                if key in seen:
                    continue
                seen.add(key)
                out.append("  %s %s → %s   %s" % (before, uom, after, component))
        return out
