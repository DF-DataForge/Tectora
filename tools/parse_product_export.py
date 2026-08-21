#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parse the Simpla product export into tectora_products seed data.

Usage:
    python3 tools/parse_product_export.py <producten_export.xlsx> [output.json]

The export holds one row per product with the purchase price, the sales price,
the unit, the supplier and the supplier's own product code. Products are
classified into the existing product categories using three signals, in order
of reliability:

1. the ``productgroup`` column, when it carries a real group;
2. keywords in the Dutch product name;
3. the supplier, whose assortment is homogeneous (Bendec sells roof edges,
   Modde Heule zinc drainage, Cintralux domes, ...).

Material families that have no matching works category become sub-categories
of "04. Opbouwwerken plat dak", so the existing category tree is reused
instead of duplicated.
"""
import json
import re
import sys
from collections import Counter, OrderedDict

import openpyxl

SHEET = "Export Simpla"
PLACEHOLDER = "opnieuw te gebruiken"

# Median sale/purchase ratio of the rows that carry both prices (see the
# analysis in the commit message); used for products without a sales price.
DEFAULT_MARKUP = 1.85

# Existing categories are referenced by their full path under the root.
CATEGORIES = OrderedDict([
    ("algemene_werken", "01. Algemene werken"),
    ("dakranden", "02. Dakranden"),
    ("veiligheid", "02. Verplichte veiligheidsvoorzieningen"),
    ("afbraak", "03. Afbraakwerken plat dak"),
    ("opbouw", "04. Opbouwwerken plat dak"),
    ("afvoer", "04. Opbouwwerken plat dak/Afvoer"),
    ("isolatie", "04. Opbouwwerken plat dak/Isolatie"),
    ("afdichting", "04. Opbouwwerken plat dak/Afdichting"),
    ("bevestiging", "04. Opbouwwerken plat dak/Bevestigingsmaterialen"),
    ("hout", "04. Opbouwwerken plat dak/Hout & plaatmateriaal"),
    ("doorvoer", "04. Opbouwwerken plat dak/Doorvoeren & schoorsteen"),
    ("gereedschap", "04. Opbouwwerken plat dak/Gereedschap & klein materiaal"),
    ("koepels", "05. Koepels"),
    ("dakramen", "06. Dakramen"),
    ("terrasvloer", "08. Terrasvloer"),
    ("andere", "12. Andere werken"),
])

# 1) productgroup -> category key
GROUP_MAP = {
    "ALGEMEEN": "algemene_werken",
    "VEILIGHEID": "veiligheid",
    "AFBRAAK": "afbraak",
    "DAKRANDEN": "dakranden",
    "AFVOER": "afvoer",
    "ISOLATIE": "isolatie",
    "AFDICHTING": "afdichting",
    "BEVESTIGINGSMATERIALEN": "bevestiging",
    "HOUT MASSIEF": "hout",
    "HOUT PLAATWERK": "hout",
    "DOORVOER": "doorvoer",
    "SCHOUW/SCHOORSTEEN": "doorvoer",
    "KOEPELS": "koepels",
    "DAKRAMEN": "dakramen",
    "DAKBELEVING": "terrasvloer",
    "GEREEDSCHAP": "gereedschap",
}
# Groups that carry no classification value.
GROUP_IGNORE = {"GEIMPORTEERDE PRODUCTEN", "DIVERSEN", "DIVERS"}

# 2) name keywords -> category key (first match wins, so order matters)
NAME_RULES = [
    ("koepels", r"koepel|skylux|lichtkoepel|lichtstraat|spindel|dakluik|daktoegang"),
    ("dakramen", r"dakra[ae]?m|velux|platdakraam"),
    ("dakranden", r"dakrand|drbs?|drbr|muurkap|kraal|daktrim|binnenhoek|buitenhoek|eindstuk"),
    ("afvoer", r"afvoer|aflooppijp|afleider|bocht|aflopen|hemelwater|tapbuis|spuwer|"
               r"noodoverloop|bladvanger|wavin|wafix|geberit|pp buis|steekmof|t-stuk|"
               r"regenwater|regenpijp|goot|zink"),
    ("isolatie", r"isolatie|utherm|pir |pir-|eps|xps|rockwool|isover|resol|"
                 r"hellingsisolatie|afschot|isolfoam"),
    ("doorvoer", r"doorvoer|schoorsteen|schouw|dakdoorvoer|ventilatie|verluchting|"
                 r"ontluchting"),
    ("afdichting", r"epdm|elevate|resitrix|roofing|bitumen|dampscherm|primer|kleefstof|"
                   r"bladlood|franaglass|franaline|aluflexx|sopra|glasvlies|pb v3|"
                   r"lijm|flashing|slab|\bkit\b|kitpatroon|mastiek|band|folie|membraan|coating|"
                   r"waterdicht|sealant|ms 500|worst"),
    ("veiligheid", r"ankerpunt|wallfix|abs-lock|valbeveiliging|balustrade|harnas|"
                   r"leeflijn|stelling|hoogwerker|weight angel|safetybull|doorfix|"
                   r"veiligheid"),
    ("bevestiging", r"slagplug|plug |plug\b|schroef|bout|nagel|bevestig|ring |"
                    r"moer|tapschroef|parco|reca |bit |zaagblad|boor"),
    ("hout", r"osb|multiplex|douglas|sls |cls |plaatmateriaal|balk|lat |latten|"
             r"spouwplank|houtvezel|triplex|profiply|geschaafd|c24"),
    ("gereedschap", r"borstel|roller|steel |telescoop|beugel|handvat|emmer|"
                    r"gereedschap|mes |spatel|brander|gasfles|kitpistool|pistool"),
    ("terrasvloer", r"terras|tegeldrager|rubbertegel|tegel|dallen|bankirai|"
                    r"vlonder|grind|kiezel|ballast|groendak|substraat|sedum"),
    ("afbraak", r"verwijderen|afbraak|uitbreken|slopen"),
    ("algemene_werken", r"werkuren|transport|werfinrichting|afvalverwerking|container|"
                        r"voorrijd|parkeervergunning|vaste kosten|werftoilet|"
                        r"klein materiaal|forfait"),
]

# 3) supplier -> category key (their assortment is homogeneous)
SUPPLIER_MAP = {
    "BENDEC": "dakranden",
    "BERTEC": "afdichting",
    "ASPHALT EQUIPMENT": "dakranden",
    "MODDE HEULE": "afvoer",
    "DEFRANCQ BOUWSPECIALITEITEN": "afvoer",
    "DEFRANCQ": "afvoer",
    "STG": "afvoer",
    "COFAPRO": "afvoer",
    "CINTRALUX": "koepels",
    "RECA BELUX": "bevestiging",
    "NOGEL": "bevestiging",
    "FRP": "afdichting",
    "VC WOOD": "hout",
    "BRUSH N ROLL": "gereedschap",
    "EYECATCHER": "veiligheid",
    "BRANDSTOFFEN TOON DEGRAUWE": "gereedschap",
}

# Unit normalisation: everything maps onto a small, clean set of UoMs. Pack
# sizes ("12 stuks", "20 BOX") are sold as one item and keep their pack info
# in the description.
UOM_RULES = [
    ("m2", r"^\d*\s*m2$|^\d*\s*m²$"),
    ("m", r"^\d*\s*(lm|ml|m)$"),
    ("hour", r"^(uur|u)$"),
    ("day", r"^per dag$|^dag$"),
    ("forfait", r"^sog$|^forfait$|^vh$"),
    ("kg", r"^kg$"),
]

# The export spells one vendor two ways.
SUPPLIER_ALIASES = {"DEFRANCQ": "DEFRANCQ BOUWSPECIALITEITEN"}

SERVICE_NAME = re.compile(
    r"^(werkuren|verwijderen|plaatsen|keuren|huur|gebruik|leveren en plaatsen van de "
    r"stelling|afvalverwerking|transport|verticaal transport|parkeervergunning|"
    r"algemene vaste kosten|extra transport)",
    re.I,
)
SERVICE_GROUPS = {"ALGEMEEN", "AFBRAAK"}


def clean(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def number(value):
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def normalise_uom(raw):
    unit = clean(raw).lower()
    for key, pattern in UOM_RULES:
        if re.match(pattern, unit):
            return key
    return "unit"


def classify(row):
    """Return (category key, signal used)."""
    group = clean(row.get("productgroup")).upper()
    if group and group not in GROUP_IGNORE:
        key = GROUP_MAP.get(group)
        if key:
            return key, "productgroup"
    name = clean(row.get("naam NL")).lower()
    for key, pattern in NAME_RULES:
        if re.search(pattern, name):
            return key, "naam"
    supplier = clean(row.get("leverancier")).upper()
    if supplier in SUPPLIER_MAP:
        return SUPPLIER_MAP[supplier], "leverancier"
    return "andere", "fallback"


def build(path):
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[SHEET] if SHEET in workbook.sheetnames else workbook.worksheets[0]
    rows = list(sheet.iter_rows(values_only=True))
    header = [clean(cell) for cell in rows[0]]

    products = OrderedDict()
    suppliers = OrderedDict()
    signals = Counter()
    categories_used = Counter()
    computed_prices = 0
    barcodes = {}

    for raw in rows[1:]:
        row = dict(zip(header, raw))
        code = clean(row.get("productcode"))
        name = clean(row.get("naam NL"))
        if not code or not name or PLACEHOLDER in name.lower():
            continue

        cost = number(row.get("aankoopprijs excl BTW"))
        sale = number(row.get("verkoopprijs excl BTW"))
        advice = number(row.get("adviesprijs voor verkoop excl BTW"))
        price_source = "verkoopprijs"
        if sale <= 0:
            if advice > 0:
                sale, price_source = advice, "adviesprijs"
            elif cost > 0:
                sale = round(cost * DEFAULT_MARKUP, 2)
                price_source = "berekend"
                computed_prices += 1
            else:
                price_source = "geen"

        category, signal = classify(row)
        signals[signal] += 1
        categories_used[category] += 1

        supplier = clean(row.get("leverancier"))
        supplier = SUPPLIER_ALIASES.get(supplier.upper(), supplier)
        if supplier:
            suppliers.setdefault(supplier, 0)
            suppliers[supplier] += 1

        unit_raw = clean(row.get("eenheid"))
        pack = clean(row.get("aantal stuks per omdoos"))
        group = clean(row.get("productgroup")).upper()
        # Every purchasable material carries a supplier; rows without one
        # are the works/labour items of the price book (services).
        is_service = not supplier

        description_bits = []
        if unit_raw and normalise_uom(unit_raw) == "unit" and unit_raw.lower() not in (
            "st", "stuk", "stuks", "stk", "1 stuk", "1stuk", "stuk(s)"
        ):
            description_bits.append("Verpakking/eenheid: %s" % unit_raw)
        if pack and pack not in ("0", "1"):
            description_bits.append("Aantal per omdoos: %s" % pack)
        brand = clean(row.get("merk"))
        if brand:
            description_bits.append("Merk: %s" % brand)

        barcode = re.sub(r"\D", "", clean(row.get("EAN-nummer")))
        if barcode and (len(barcode) < 8 or barcode in barcodes):
            barcode = ""  # keep barcodes unique and plausible
        if barcode:
            barcodes[barcode] = code

        products[code] = {
            "code": code,
            "name": name[:180],
            "category": category,
            "category_signal": signal,
            "uom": normalise_uom(unit_raw),
            "unit_raw": unit_raw,
            "list_price": round(sale, 2),
            "standard_price": round(cost, 2),
            "price_source": price_source,
            "supplier": supplier,
            "supplier_code": clean(row.get("productcode leverancier")),
            "is_service": is_service,
            "barcode": barcode,
            "weight": number(row.get("gewicht")),
            "description": " · ".join(description_bits),
            "tags": (["Prijs berekend"] if price_source == "berekend" else [])
            + (["Prijs op aanvraag"] if sale <= 0 else []),
        }

    return {
        "categories": [
            {"key": key, "path": path_} for key, path_ in CATEGORIES.items()
        ],
        "root_category": "Dakwerken",
        "suppliers": sorted(suppliers),
        "markup": DEFAULT_MARKUP,
        "products": list(products.values()),
        "stats": {
            "products": len(products),
            "computed_prices": computed_prices,
            "signals": dict(signals),
            "per_category": {
                CATEGORIES[key]: count for key, count in categories_used.most_common()
            },
            "suppliers": suppliers,
        },
    }


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    output = (
        sys.argv[2] if len(sys.argv) > 2
        else "tectora_products/data/product_catalog.json"
    )
    data = build(sys.argv[1])
    with open(output, "w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=1)
    stats = data["stats"]
    print("%s products -> %s" % (stats["products"], output))
    print("classification signals:", stats["signals"])
    print("computed sales prices (%.2f x aankoop): %s"
          % (data["markup"], stats["computed_prices"]))
    print("suppliers:", len(data["suppliers"]))
    print("per category:")
    for name, count in stats["per_category"].items():
        print("   %5d  %s" % (count, name))


if __name__ == "__main__":
    main()
