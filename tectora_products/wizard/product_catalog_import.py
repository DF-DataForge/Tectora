# -*- coding: utf-8 -*-
import base64
import io
import logging

from odoo import _, fields, models
from odoo.exceptions import UserError

from ..models import catalog_rules

_logger = logging.getLogger(__name__)


class ProductCatalogImport(models.TransientModel):
    _name = "tectora.product.catalog.import"
    _description = "Catalogus importeren"

    file = fields.Binary(string="Bestand (.xlsx)", required=True, attachment=False)
    filename = fields.Char(string="Bestandsnaam")
    sheet_name = fields.Char(
        string="Tabblad",
        help="Laat leeg om het eerste tabblad (of 'Export Simpla') te nemen.",
    )
    category_mode = fields.Selection(
        [
            ("auto", "Automatisch indelen in de bestaande categorieën"),
            ("column", "Uit de kolom 'productcategorie' in het bestand"),
        ],
        string="Productcategorie",
        default="auto",
        required=True,
        help="Automatisch: de productgroep, de naam en de leverancier bepalen "
        "de bestaande categorie. Uit de kolom: de categorie in het bestand "
        "wordt letterlijk gebruikt (en indien nodig aangemaakt); een pad met "
        "'/' maakt subcategorieën.",
    )
    type_mode = fields.Selection(
        [
            ("auto", "Automatisch: met leverancier = grondstof, zonder = dienst"),
            ("column", "Uit een type-kolom in het bestand"),
            ("service", "Alles als verkoopproduct (dienst)"),
            ("goods", "Alles als grondstof (voorraadproduct)"),
        ],
        string="Soort product",
        default="auto",
        required=True,
        help="Verkoopproducten worden diensten (verkoopbaar, niet in "
        "voorraad); grondstoffen worden voorraadproducten (aankoopbaar, met "
        "voorraadbeheer).",
    )
    markup = fields.Float(
        string="Marge-coëfficiënt",
        default=catalog_rules.DEFAULT_MARKUP,
        help="Gebruikt als verkoopprijs voor producten die er geen hebben: "
        "aankoopprijs × deze coëfficiënt. 1,85 is de mediaan van de "
        "bestaande catalogus.",
    )
    sale_ok_goods = fields.Boolean(
        string="Grondstoffen ook verkoopbaar",
        default=False,
        help="Standaard zijn grondstoffen enkel aankoopbaar: enkel de "
        "verkoopproducten (diensten) verschijnen dan in offertes en in de "
        "productkiezer van de tekening.",
    )
    archive_price_book = fields.Boolean(
        string="Oude prijsboekproducten archiveren",
        default=False,
        help="Archiveert de producten met een 'DAK-' referentie. Ze worden "
        "nooit verwijderd, zodat bestaande offertes intact blijven.",
    )
    state = fields.Selection(
        [("draft", "Concept"), ("done", "Uitgevoerd")], default="draft"
    )
    summary = fields.Text(string="Resultaat", readonly=True)

    # ------------------------------------------------------------------ import
    def _read_rows(self):
        """Return (header, rows) of the uploaded workbook."""
        self.ensure_one()
        try:
            import openpyxl  # Odoo ships openpyxl (see requirements.txt)
        except ImportError:
            raise UserError(
                _("De Python-bibliotheek openpyxl is niet beschikbaar op de server.")
            )
        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(base64.b64decode(self.file)),
                data_only=True,
                read_only=True,
            )
        except Exception as error:
            raise UserError(
                _("Het bestand kon niet gelezen worden als .xlsx-bestand (%s).",
                  error)
            )
        name = (self.sheet_name or "").strip()
        if name and name in workbook.sheetnames:
            sheet = workbook[name]
        elif name:
            raise UserError(
                _("Tabblad '%(sheet)s' niet gevonden. Beschikbaar: %(all)s",
                  sheet=name, all=", ".join(workbook.sheetnames))
            )
        elif "Export Simpla" in workbook.sheetnames:
            sheet = workbook["Export Simpla"]
        else:
            sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            raise UserError(_("Het tabblad bevat geen gegevens."))
        return rows[0], rows[1:]

    def action_import(self):
        self.ensure_one()
        header, rows = self._read_rows()
        try:
            entries, stats = catalog_rules.parse_rows(
                header,
                rows,
                {
                    "category_mode": self.category_mode,
                    "type_mode": self.type_mode,
                    "markup": self.markup or catalog_rules.DEFAULT_MARKUP,
                },
            )
        except ValueError as error:
            raise UserError(str(error))
        if not entries:
            raise UserError(_("Geen producten gevonden in het bestand."))

        counters = self.env["product.template"]._tectora_apply_catalog(
            entries, {"sale_ok_goods": self.sale_ok_goods}
        )
        archived = self.env["product.template"]._tectora_archive_price_book(
            self.archive_price_book
        )
        _logger.info(
            "tectora_products: wizard import %s (%s)", self.filename, counters
        )

        lines = [
            _("Bestand: %s", self.filename or "-"),
            _("%(new)s nieuwe producten, %(upd)s bijgewerkt",
              new=counters["created"], upd=counters["updated"]),
            _("%(svc)s verkoopproducten (diensten), %(goods)s grondstoffen "
              "(voorraadproducten)",
              svc=stats["services"], goods=stats["goods"]),
            _("%s leverancierslijnen (aankoopprijs + leverancierscode)",
              counters["vendor_lines"]),
            _("%(n)s verkoopprijzen berekend (aankoop × %(factor).2f)",
              n=stats["computed_prices"],
              factor=self.markup or catalog_rules.DEFAULT_MARKUP),
        ]
        if stats["signals"]:
            lines.append(
                _("Categorie bepaald via: %s")
                % ", ".join(
                    "%s: %s" % (signal, count)
                    for signal, count in sorted(stats["signals"].items())
                )
            )
        if stats["without_code"]:
            lines.append(
                _("%s producten zonder interne referentie (gematcht op naam "
                  "binnen de categorie)") % stats["without_code"]
            )
        if stats["duplicates"]:
            lines.append(
                _("%s dubbele referenties overgeslagen") % stats["duplicates"]
            )
        if stats["skipped"]:
            lines.append(_("%s lege of placeholder-rijen overgeslagen")
                         % stats["skipped"])
        if archived:
            lines.append(_("%s oude prijsboekproducten gearchiveerd") % archived)

        self.write({"state": "done", "summary": "\n".join(lines)})
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
            "context": self.env.context,
        }

